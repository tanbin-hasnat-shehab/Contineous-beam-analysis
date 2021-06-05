from tkinter import *
from PIL import ImageTk,Image
from tkinter import ttk
import math
import os
import numpy as np
from openpyxl.workbook import *
from openpyxl import load_workbook




'''os.startfile("deleting_catched_data.exe")'''
try:
    os.remove('span_lengths.txt')
    
    
    
except:
    pass

try:
   
    os.remove('inertia.txt')
    os.remove('all_datos.xlsx')
   
    
except:
    pass

try:
    
    os.remove('point_load_distances2.txt')
    
    
except:
    pass


try:
    
    os.remove('point_loads.txt')
    
    
except:
    pass


try:
    
    os.remove('distributed_loads.txt')
    
    
    
except:
    pass


try:
    
    os.remove('fixed_end_moments.txt')
    
    
except:
    pass




  



root=Tk()

'''root.geometry("%dx%d+0+0"%(root.winfo_screenwidth(),root.winfo_screenheight()))'''
root.geometry("1000x400")

''''
root er upor main frame create krte hbe
mainframe er upor canvas
main frame er upor scrollbar+configuration
canvas er upor secondary frame create krte hbe jetate sob kaj krte hbe
create korar por canvas er sathe attach kore dite hbe


'''




###creating scrollable canvas
#1st frame
main_frame=Frame(root)
main_frame.pack(fill=BOTH,expand=1)



#canvas create
my_canvas=Canvas(main_frame)
my_canvas.pack(side=BOTTOM,fill=BOTH,expand=1)



#create scroll bar
my_scrollbar=ttk.Scrollbar(main_frame,orient=HORIZONTAL,command=my_canvas.xview)
my_scrollbar.pack(side=LEFT,fill=X)

my_scrollbar1=ttk.Scrollbar(main_frame,orient=VERTICAL,command=my_canvas.yview)
my_scrollbar1.pack(side=RIGHT,fill=Y)




#config canvas
my_canvas.configure(xscrollcommand=my_scrollbar.set)
my_canvas.configure(yscrollcommand=my_scrollbar1.set)
my_canvas.bind('<Configure>',lambda e: my_canvas.configure(scrollregion=my_canvas.bbox("all")))




#another frame create
second_frame=Frame(my_canvas)





#add the new frame to the canvas
my_canvas.create_window((0,0),window=second_frame,anchor="nw")

#scrolling ends here  ------------------------new root is "second_frame"-----------------------------




mid_img=ImageTk.PhotoImage(Image.open("beams\\7.png"))
img1=ImageTk.PhotoImage(Image.open("beams\\1.png"))
img2=ImageTk.PhotoImage(Image.open("beams\\2.png"))
img3=ImageTk.PhotoImage(Image.open("beams\\3.png"))
img4=ImageTk.PhotoImage(Image.open("beams\\4.png"))
img5=ImageTk.PhotoImage(Image.open("beams\\5.png"))
img6=ImageTk.PhotoImage(Image.open("beams\\6.png"))
P=ImageTk.PhotoImage(Image.open("beams\\P.png"))


###########fn 2 4 5 6  



    

def fn2():
    
    
    def fn4():
        
        def fn5():
            all_length=[]
            index=0
            for i in range(0,int(span)):
                length=0
                for j in range(0,int(cum_load_dist[i].get())+1):
                    length=length+float(only_dist[j+index].get())
                    if j==int(cum_load_dist[i].get()):
                        index=index+int(cum_load_dist[i].get())+1
                all_length.append(length)
            lengths_pointer=open("span_lengths.txt","a")    
            for i in all_length:
                lengths_pointer.write(str(i)+"\n")
            lengths_pointer.close()
            
            
            
            wb=Workbook()
            omegas_sheet=wb.create_sheet('omegas')
            span_lengths_sheet=wb.create_sheet('span_lengths') 

            inertias_sheet=wb.create_sheet('inertias')
            point_loads_sheet=wb.create_sheet('poin_loads')
            distances_sheet=wb.create_sheet('distances_between_point_loads')
            fem_sheet=wb.create_sheet('fixed_end_moments')
            
            for i in range(0,len(all_length)):
                span_lengths_sheet.cell(column=1,row=i+1,value=all_length[i])
            
            inertia=[]
            for i in range(0,len(beam_width)):
                xx=float(beam_width[i].get())*pow(float(beam_height[i].get()),3)/12
                inertia.append(xx)
                
            minimum=min(inertia)
            Inertia=[]
            for i in inertia:
                yy=i/minimum
                Inertia.append(yy)
            
            
            
                
            
                
                
                
                
                
                
                
                
            inertia_pointer=open("inertia.txt","a")    
            for i in Inertia:
                inertia_pointer.write(str(i)+"\n")
            inertia_pointer.close() 

            for i in range(0,len(Inertia)):
                inertias_sheet.cell(column=1,row=i+1,value=Inertia[i])                 
###########fn 2 =1 tabb    fn4=2 tabb   fn 5=3tabb    fn6=4tabb  
#########fn5
                        
            print(type(all_length[0]))
            
            print(type(p_load[0].get()))
            print(type(cum_load_dist[0].get()))
            print(type(only_dist[0].get()))
            
            def fn6(first_limit,second_limit,full_length,omega,indice):
 ############fn6 starts
                               
                '''print(first_limit)
                print(second_limit)
                print(full_length)
                print("\n\n")'''
                omegas_sheet.cell(row=indice+1,column=1,value=omega)
                my_load=[]
                distances=[]
                for i in range(first_limit,second_limit):
                    my_load.append(float(p_load[i].get()))
                    
                p_pointer=open("point_loads.txt","a")
                p_pointer.write(str(indice))
                for i in my_load:
                    p_pointer.write("\n\t"+str(i))
                p_pointer.write("\n\n")
                p_pointer.close()
                
                for i in range(0,len(my_load)):
                    
                    point_loads_sheet.cell(column=indice+1,row=i+1,value=my_load[i])
                    
                 
                print("\n\n")
                if first_limit==0:
                    for i in range(0,second_limit+1):
                        distances.append(float(only_dist[i].get()))
                        
                else:
                    for i in range(first_limit+indice,second_limit+indice+1):
                        distances.append(float(only_dist[i].get()))
                
                
                distances_pointer=open("point_load_distances2.txt","a")
                distances_pointer.write(str(indice))
                for i in distances:
                    distances_pointer.write("\n\t"+str(i))
                distances_pointer.write("\n\n")
                distances_pointer.close()
                
                
                for i in range(0,len(my_load)+1):
                    
                    distances_sheet.cell(column=indice+1,row=i+1,value=distances[i])
                
                
                omega_pointer=open("distributed_loads.txt","a")
                omega_pointer.write(str(omega)+"\n")
                omega_pointer.close()
                
                
                
                
                
                
                '''if first_limit==0:
                    for i in range(0,second_limit+1):
                        distances_pointer.write(str(distances[i]))
                        
                        
                else:
                    for i in range(first_limit+indice,second_limit+indice+1):
                        distances_pointer.write(str(distances[i]))'''
                                
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                
                fem1=omega*full_length*full_length/12+(my_load[0]*math.pow((distances[0]),1)*math.pow((full_length-(distances[0])),2)/(math.pow(full_length,2)))+00000+(my_load[1]*math.pow((distances[0]+distances[1]),1)*math.pow((full_length-(distances[0]+distances[1])),2)/(math.pow(full_length,2)))+00000+(my_load[2]*math.pow((distances[0]+distances[1]+distances[2]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2])),2)/(math.pow(full_length,2)))+00000+(my_load[3]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3])),2)/(math.pow(full_length,2)))+00000+(my_load[4]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4])),2)/(math.pow(full_length,2)))+00000+(my_load[5]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5])),2)/(math.pow(full_length,2)))+00000+(my_load[6]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6])),2)/(math.pow(full_length,2)))+00000+(my_load[7]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7])),2)/(math.pow(full_length,2)))+00000+(my_load[8]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8])),2)/(math.pow(full_length,2)))+00000+(my_load[9]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9])),2)/(math.pow(full_length,2)))+00000+(my_load[10]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]+distances[10]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]+distances[10])),2)/(math.pow(full_length,2)))
                
                fem2=-(omega*full_length*full_length/12+(my_load[0]*math.pow((distances[0]),2)*(full_length-(distances[0]))/(math.pow(full_length,2)))+00000+(my_load[1]*math.pow((distances[0]+distances[1]),2)*(full_length-(distances[0]+distances[1]))/(math.pow(full_length,2)))+00000+(my_load[2]*math.pow((distances[0]+distances[1]+distances[2]),2)*(full_length-(distances[0]+distances[1]+distances[2]))/(math.pow(full_length,2)))+00000+(my_load[3]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]))/(math.pow(full_length,2)))+00000+(my_load[4]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]))/(math.pow(full_length,2)))+00000+(my_load[5]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]))/(math.pow(full_length,2)))+00000+(my_load[6]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]))/(math.pow(full_length,2)))+00000+(my_load[7]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]))/(math.pow(full_length,2)))+00000+(my_load[8]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]))/(math.pow(full_length,2)))+00000+(my_load[9]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]))/(math.pow(full_length,2)))+00000+(my_load[10]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]+distances[10]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]+distances[10]))/(math.pow(full_length,2))))
                
                
                
                '''lengths_pointer=open("span_lengths.txt","a")    
                for i in all_length:
                    lengths_pointer.write(str(i))
                    lengths_pointer.close()'''
                    
                
                    
                    
                print("length = ",full_length)
                print("omega = ",omega)
                
                fem_pointer=open("fixed_end_moments.txt","a")
                fem_pointer.write(str(indice)+"\n\t"+str(fem1)+"\n\t"+str(fem2)+"\n")
                fem_pointer.close()
                
                    
                fem_sheet.cell(column=1,row=2*indice+1,value=fem1)
                fem_sheet.cell(column=1,row=2*indice+2,value=fem2)
                wb.save('all_datos.xlsx')
                
                    
                
                
                
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                    
                print(fem1)
                print(fem2)
                del distances
                
                
 ############fn6 ends 
                
            global indexer
            indexer=0
            for i in range(0,int(span)):
                incc=i
                bttn=Button(second_frame,text="load",width=5,command=fn6(indexer,indexer+int(cum_load_dist[i].get()),all_length[i],float(distributed_loads[i].get()),incc))
                bttn.place(x=257.5+515*i,y=130)
                indexer=indexer+int(cum_load_dist[i].get())
            
                        
            
            
            
                        
                        
               
###########fn 2 =1 tabb    fn4=2 tabb   fn 5=3tabb    fn6=4tabb                
            
        
        btncal=Button(second_frame,text="load data",width=30,command=fn5)
        btncal.grid(row=8,column=0)
        
        
        
        
###########fn 2 =1 tabb    fn4=2 tabb   fn 5=3tabb    fn6=4tabb      
        
        p_load=[]
        a=0
        n=0
        for i in range(0,len(cum_load_dist)):
           
            for j in range(0,int(cum_load_dist[i].get())):
                P_load=Entry(second_frame,width=5,bg="#99ccff")
                P_load.place(x=a+515/(int(cum_load_dist[i].get())+1),y=183)
                p_load.append(P_load)
                
                point_load_img=Label(second_frame,image=P)
                point_load_img.place(x=a+515/(int(cum_load_dist[i].get())+1),y=203)
                
                a=a+515/(int(cum_load_dist[i].get())+1)
                if j==int(cum_load_dist[i].get())-1:
                    n=n+1
                    a=515*n
                            
        only_dist=[]            
        b=0
        m=0            
        for i in range(0,len(cum_load_dist)):
            for j in range(0,int(cum_load_dist[i].get())+1):
                Only_dist=Entry(second_frame,borderwidth=4,width=3,bg="#99ccff")
                Only_dist.place(x=b+(515/(int(cum_load_dist[i].get())+1))/2*(2*j+1),y=315)
                only_dist.append(Only_dist)
                
                
                if j==int(cum_load_dist[i].get()):
                    m=m+1
                    b=515*m
                
            
            
###########fn 2 =1 tabb    fn4=2 tabb   fn 5=3tabb    fn6=4tabb        
        
        
        

    
    pl=Label(second_frame,height=10).grid(row=5,column=0,columnspan=int(span))
    
    
    cum_load_dist=[]
    list_in_mid_span=[]
    var_img=Variable
    in_spn=[]
    lbl=[]
    for i in range(0,int(span)):
        
        if i==0 and left=='fixed':
            
            lbl=Label(second_frame,image=img1,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)'''
            
            
        if i==0 and left=='hinge':
            lbl=Label(second_frame,image=img2,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)'''
        
        
        if i==0 and left=='roller':
            lbl=Label(second_frame,image=img3,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)'''
        
        
        if i==int(span)-1 and right=='fixed':
            lbl=Label(second_frame,image=img4,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)'''
            
            
        if i==int(span)-1 and right=='hinge':
            lbl=Label(second_frame,image=img5,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)''' 
            
            
        if i==int(span)-1 and right=='roller':
            lbl=Label(second_frame,image=img6,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)'''
        
        if i>0 and i!=int(span)-1:
            lbl=Label(second_frame,image=mid_img,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)''' 
        Cum_load_dist=Entry(second_frame,width=10)
        Cum_load_dist.grid(row=3,column=i)
        cum_load_dist.append(Cum_load_dist)
        pick_btn=Button(second_frame,text=i,command=fn4)
        pick_btn.grid(row=4,column=i)
    distributed_loads=[]
    beam_width=[]
    beam_height=[]
        
    for i in range(0,int(span)):
        distributed_input=Entry(second_frame,width=5)
        distributed_input.place(x=257.5+515*i,y=275)
        distributed_loads.append(distributed_input)
        
        beam_w=Label(second_frame,text="B=",width=3)
        beam_w.place(x=257.5-100+515*i,y=340)
        
        
        b_width=Entry(second_frame,bg="#33FFA8",width=5)
        b_width.place(x=257.5-75+515*i,y=340)
        beam_width.append(b_width)
        
        beam_h=Label(second_frame,text="D=",width=3)
        beam_h.place(x=257.5+515*i,y=340)
        
        
        b_ht=Entry(second_frame,bg="#33FFA8",width=5)
        b_ht.place(x=257.5+25+515*i,y=340)
        beam_height.append(b_ht)
        
        
        
    
        
        
        
        
###########fn 2 =1 tabb    fn4=2 tabb   fn 5=3tabb    fn6=4tabb                    
        
       
    '''load_span_btn=Button(second_frame,text="input loads",width=20,command=fn4).grid(row=8,column=0)'''
    
    



def fn1():
    try:
        global span,left,right
        span=float(Ent1.get())
        
        left=(clicked.get())
        right=(clicked1.get())
        fn2()
        
    except ValueError:
        pass
    
    
            

label1=Label(second_frame,text="no. of span",width=10).grid(row=0,column=0)
Ent1=Entry(second_frame,width=10,borderwidth=5)

Ent1.grid(row=0,column=1)




clicked=StringVar()
clicked.set("select")

clicked1=StringVar()
clicked1.set("select")

label2=Label(second_frame,text="Left end",width=50).grid(row=1,column=0)

drop1=OptionMenu(second_frame,clicked,"fixed","hinge","roller").grid(row=1,column=1)

label3=Label(second_frame,text="Right end",width=50).grid(row=2,column=0)

drop2=OptionMenu(second_frame,clicked1,"fixed","hinge","roller").grid(row=2,column=1)

btn1=Button(second_frame,text="ok",width=30,command=fn1).grid(row=2,column=2)
'''btn2=Button(second_frame,text="ok",width=30,command=second_frame.close).grid(row=2,column=3)'''








    
    


















root.mainloop()
