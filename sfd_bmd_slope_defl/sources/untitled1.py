import numpy as np
import sympy as sy
from sympy import solve_linear_system
from sympy import*
from openpyxl.workbook import *
from openpyxl import load_workbook
import matplotlib.pyplot as plt
%matplotlib inline


sy.init_printing()

wb=load_workbook('all_datos.xlsx')
omega_sheet=wb['omegas']
span_lengths_sheet=wb['span_lengths']
moments_sheet=wb['moments_after_calculation']
point_load_sheet=wb['poin_loads']
distances_sheet=wb['distances_between_point_loads']


total_spans=(len(span_lengths_sheet['a']))

def fn1(span_number,b,cum_dist):
    print("---------------------------------")
    
    p=[]
    distances=[]
    for i in range(1,100):
        p.append(point_load_sheet.cell(row=i,column=span_number).value)
        if point_load_sheet.cell(row=i+1,column=span_number).value==None:
            break
    for i in range(1,100):
        distances.append(distances_sheet.cell(row=i,column=span_number).value)
        if distances_sheet.cell(row=i+1,column=span_number).value==None:
            break    
    
    print("\ndistances are")
    for i in distances:
        print(i)
    distant=[]
    xx=0
    for i in range(0,len(p)):
        xx=xx+distances[i]
        distant.append(xx)
        
        
        
        
        
    print("\n point loads are")
    for i in p:
        print(i)
    moment_left=-1*float((moments_sheet.cell(row=b,column=1).value))
    moment_right=-1*float((moments_sheet.cell(row=b+1,column=1).value))
    omega=omega_sheet.cell(row=span_number,column=1).value
    L=span_lengths_sheet.cell(row=span_number,column=1).value
    
    print(f'moment left = {moment_left}  moment right = {moment_right}  omega= {omega}  span length= {L}')
    
     
    print('\n')
    arbit_sum=0
    for i in range(0,len(p)):
        arbit=p[i]*distant[i]
        arbit_sum=arbit_sum+arbit
    R_left=omega*L+sum(p)-(moment_left+moment_right+arbit_sum+omega*L**2/2)/L
    print("R left is    = ",R_left)
    R_right=(moment_left+moment_right+arbit_sum+omega*L**2/2)/L
    ###############data points for sfd
    
    
    
    total_points=len(p)
    print("total points  = ",total_points)
    
    point_y=[]
    point_y.append(0)
    point_y.append(R_left)
    
    
    z=0
    for i in range(0,total_points):
        if z==0:
            point_y.append(R_left-omega*distances[i])
            point_y.append(point_y[2]-p[i])
        else:
            point_y.append(point_y[z+1]-omega*distances[i])
            point_y.append(point_y[z+2]-p[i])
        z+=2
    point_y.append(-1*R_right)
    point_y.append(0)
    
        
    point_x=[]
    point_x.append(0+cum_dist)
    point_x.append(0+cum_dist)
    for i in range(0,len(distant)):
        point_x.append(cum_dist+distant[i])
        point_x.append(cum_dist+distant[i])
    point_x.append(L+cum_dist)
    point_x.append(L+cum_dist)
    
    
    for i in point_y:
        print("y = ",i)
        
    for i in point_x:
        print("x = ",i)
    ##################
    #moments points starts here
    
    
    
    
    
    print("\n\ndistant")
    for i in distant:
        print(i)
        
        
    mnt_right=-1*moment_right
    m_point_y=[]
    
    m_point_y.append(moment_left)
    for i in range(1,L):
        c=0
        for j in distant:
            if j<i:
                c+=1
        if c==0:
            m_y=moment_left+R_left*i-(omega*i*i)/2
            m_point_y.append(m_y)
        else:
            suml=0
            for k in range(0,c):
                suml=suml+(p[k])*(i-distant[k])
            m_y=moment_left+R_left*i-(omega*i*i)/2-suml
            m_point_y.append(m_y)
    m_point_y.append(mnt_right)
    
    for i in m_point_y:
        print(i)
    
    print("total then x es",len(m_point_y))
    m_point_x=[]
  
    for i in range(0,len(m_point_y)):
        m_point_x.append(i+cum_dist)
    for i in m_point_x:
        print(i)
####--------------------------     
    for i in m_point_x:
        moments_x_d.append(i)
    
    for i in m_point_y:
        moments_y_d.append(i)
    for i in point_x:
        sfd_x_d.append(i)
    for i in point_y:
        sfd_y_d.append(i)
   
    
    
    
    
    
    
    
    
    #############
    
sfd_bmd_sheet=wb.create_sheet('drawings')   
lengths=[]
for i in range(1,len(span_lengths_sheet['a'])+1 ):
    lengths.append(span_lengths_sheet.cell(row=i,column=1).value)
    
###-----------
moments_x_d=[]
moments_y_d=[]
sfd_x_d=[]
sfd_y_d=[]

###-------------
z=0
bb=1
su=0

for i in range(1,total_spans+1):
    fn1(i,bb,su)
    su=su+lengths[i-1]
    
    
    bb+=2

sfd_bmd_sheet.cell(column=4,row=1,value=0)
for i in range(0,len(moments_y_d)):
    sfd_bmd_sheet.cell(column=4,row=i+2,value=moments_y_d[i])
sfd_bmd_sheet.cell(column=4,row=2+len(moments_y_d),value=0)



sfd_bmd_sheet.cell(column=3,row=1,value=0)
for i in range(0,len(moments_x_d)):
    sfd_bmd_sheet.cell(column=3,row=i+2,value=moments_x_d[i])
sfd_bmd_sheet.cell(column=3,row=2+len(moments_x_d),value=moments_x_d[len(moments_x_d)-1])



for i in range(0,len(sfd_x_d)):
    sfd_bmd_sheet.cell(column=1,row=i+1,value=sfd_x_d[i])

for i in range(0,len(sfd_y_d)):
    sfd_bmd_sheet.cell(column=2,row=i+1,value=sfd_y_d[i])
    
wb.save('all_datos_and_dwg.xlsx')