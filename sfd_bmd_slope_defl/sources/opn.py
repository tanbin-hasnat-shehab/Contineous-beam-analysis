from openpyxl.workbook import *
from openpyxl import load_workbook



wb=load_workbook('datos2.xlsx')
sheet1=wb.active

'''
x=[1,2,3,4,4444,4,0,0,25,25,2,52,35,2,4,1,22,23,33,6552,5,22,5,2,2]
for i in range(1,len(x)):
    sheet1.cell(row=i,column=1,value=x[i])
    '''
col_a=sheet1['a']    
x=[]
for i in range(5,len(col_a)):
    x.append(col_a[i].value)
    
x.append(5454545)   
for i in x:
    print(i)





