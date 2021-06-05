from tkinter import *
from PIL import ImageTk,Image
from tkinter import ttk
import math
import os
import numpy as np
from openpyxl.workbook import *
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import sympy as sy
from sympy import*




try:
    os.remove('inertia.txt')
    os.remove('all_datos.xlsx')    
except:
    pass


root=Tk()
root.title("shear force bending moment diag n calculation")

'''root.geometry("%dx%d+0+0"%(root.winfo_screenwidth(),root.winfo_screenheight()))'''
root.geometry("1000x400")

''''
root er upor main frame create krte hbe
mainframe er upor canvas
main frame er upor scrollbar+configuration
canvas er upor secondary frame create krte hbe jetate sob kaj krte hbe
create korar por canvas er sathe attach kore dite hbe


'''




###creating scrollable canvas
#1st frame
main_frame=Frame(root)
main_frame.pack(fill=BOTH,expand=1)



#canvas create
my_canvas=Canvas(main_frame)
my_canvas.pack(side=BOTTOM,fill=BOTH,expand=1)



#create scroll bar
my_scrollbar=ttk.Scrollbar(main_frame,orient=HORIZONTAL,command=my_canvas.xview)
my_scrollbar.pack(side=LEFT,fill=X)

my_scrollbar1=ttk.Scrollbar(main_frame,orient=VERTICAL,command=my_canvas.yview)
my_scrollbar1.pack(side=RIGHT,fill=Y)




#config canvas
my_canvas.configure(xscrollcommand=my_scrollbar.set)
my_canvas.configure(yscrollcommand=my_scrollbar1.set)
#my_canvas.bind('<Configure>',lambda e: my_canvas.configure(scrollregion=my_canvas.bbox("all")))




#another frame create
second_frame=Frame(my_canvas)





#add the new frame to the canvas
my_canvas.create_window((0,0),window=second_frame,anchor="nw")

#scrolling ends here  ------------------------new root is "second_frame"-----------------------------




mid_img=ImageTk.PhotoImage(Image.open("beams\\7.png"))
img1=ImageTk.PhotoImage(Image.open("beams\\1.png"))
img2=ImageTk.PhotoImage(Image.open("beams\\2.png"))
img3=ImageTk.PhotoImage(Image.open("beams\\3.png"))
img4=ImageTk.PhotoImage(Image.open("beams\\4.png"))
img5=ImageTk.PhotoImage(Image.open("beams\\5.png"))
img6=ImageTk.PhotoImage(Image.open("beams\\6.png"))
P=ImageTk.PhotoImage(Image.open("beams\\P.png"))


###########fn 2 4 5 6  



    

def fn2():
    
    
    def fn4():
        
        def fn5():
            all_length=[]
            index=0
            for i in range(0,int(span)):
                length=0
                for j in range(0,int(cum_load_dist[i].get())+1):
                    length=length+float(only_dist[j+index].get())
                    if j==int(cum_load_dist[i].get()):
                        index=index+int(cum_load_dist[i].get())+1
                all_length.append(length)
            lengths_pointer=open("span_lengths.txt","a")    
            for i in all_length:
                lengths_pointer.write(str(i)+"\n")
            lengths_pointer.close()
            
            
            
            wb=Workbook()
            omegas_sheet=wb.create_sheet('omegas')
            span_lengths_sheet=wb.create_sheet('span_lengths') 

            inertias_sheet=wb.create_sheet('inertias')
            point_loads_sheet=wb.create_sheet('poin_loads')
            distances_sheet=wb.create_sheet('distances_between_point_loads')
            fem_sheet=wb.create_sheet('fixed_end_moments')
            
            for i in range(0,len(all_length)):
                span_lengths_sheet.cell(column=1,row=i+1,value=all_length[i])
            span_lengths_sheet.cell(column=2,row=1,value=left)
            span_lengths_sheet.cell(column=2,row=2,value=right)
            inertia=[]
            for i in range(0,len(beam_width)):
                xx=float(beam_width[i].get())*pow(float(beam_height[i].get()),3)/12
                inertia.append(xx)
                
            minimum=min(inertia)
            Inertia=[]
            for i in inertia:
                yy=i/minimum
                Inertia.append(yy)
            
            
            
                
            
                
                
                
                
                
                
                
                
            inertia_pointer=open("inertia.txt","a")    
            for i in Inertia:
                inertia_pointer.write(str(i)+"\n")
            inertia_pointer.close() 

            for i in range(0,len(Inertia)):
                inertias_sheet.cell(column=1,row=i+1,value=Inertia[i])                 
###########fn 2 =1 tabb    fn4=2 tabb   fn 5=3tabb    fn6=4tabb  
#########fn5
                        
            print(type(all_length[0]))
            
            print(type(p_load[0].get()))
            print(type(cum_load_dist[0].get()))
            print(type(only_dist[0].get()))
            
            def fn6(first_limit,second_limit,full_length,omega,indice):
 ############fn6 starts
                               
                '''print(first_limit)
                print(second_limit)
                print(full_length)
                print("\n\n")'''
                omegas_sheet.cell(row=indice+1,column=1,value=omega)
                my_load=[]
                distances=[]
                for i in range(first_limit,second_limit):
                    my_load.append(float(p_load[i].get()))
                    
                p_pointer=open("point_loads.txt","a")
                p_pointer.write(str(indice))
                for i in my_load:
                    p_pointer.write("\n\t"+str(i))
                p_pointer.write("\n\n")
                p_pointer.close()
                
                for i in range(0,len(my_load)):
                    
                    point_loads_sheet.cell(column=indice+1,row=i+1,value=my_load[i])
                    
                 
                print("\n\n")
                if first_limit==0:
                    for i in range(0,second_limit+1):
                        distances.append(float(only_dist[i].get()))
                        
                else:
                    for i in range(first_limit+indice,second_limit+indice+1):
                        distances.append(float(only_dist[i].get()))
                
                
                distances_pointer=open("point_load_distances2.txt","a")
                distances_pointer.write(str(indice))
                for i in distances:
                    distances_pointer.write("\n\t"+str(i))
                distances_pointer.write("\n\n")
                distances_pointer.close()
                
                
                for i in range(0,len(my_load)+1):
                    
                    distances_sheet.cell(column=indice+1,row=i+1,value=distances[i])
                
                
                omega_pointer=open("distributed_loads.txt","a")
                omega_pointer.write(str(omega)+"\n")
                omega_pointer.close()
                
                
                
                
                
                
                '''if first_limit==0:
                    for i in range(0,second_limit+1):
                        distances_pointer.write(str(distances[i]))
                        
                        
                else:
                    for i in range(first_limit+indice,second_limit+indice+1):
                        distances_pointer.write(str(distances[i]))'''
                                
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                distances.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                my_load.append(0)
                
                fem1=omega*full_length*full_length/12+(my_load[0]*math.pow((distances[0]),1)*math.pow((full_length-(distances[0])),2)/(math.pow(full_length,2)))+00000+(my_load[1]*math.pow((distances[0]+distances[1]),1)*math.pow((full_length-(distances[0]+distances[1])),2)/(math.pow(full_length,2)))+00000+(my_load[2]*math.pow((distances[0]+distances[1]+distances[2]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2])),2)/(math.pow(full_length,2)))+00000+(my_load[3]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3])),2)/(math.pow(full_length,2)))+00000+(my_load[4]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4])),2)/(math.pow(full_length,2)))+00000+(my_load[5]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5])),2)/(math.pow(full_length,2)))+00000+(my_load[6]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6])),2)/(math.pow(full_length,2)))+00000+(my_load[7]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7])),2)/(math.pow(full_length,2)))+00000+(my_load[8]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8])),2)/(math.pow(full_length,2)))+00000+(my_load[9]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9])),2)/(math.pow(full_length,2)))+00000+(my_load[10]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]+distances[10]),1)*math.pow((full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]+distances[10])),2)/(math.pow(full_length,2)))
                
                fem2=-(omega*full_length*full_length/12+(my_load[0]*math.pow((distances[0]),2)*(full_length-(distances[0]))/(math.pow(full_length,2)))+00000+(my_load[1]*math.pow((distances[0]+distances[1]),2)*(full_length-(distances[0]+distances[1]))/(math.pow(full_length,2)))+00000+(my_load[2]*math.pow((distances[0]+distances[1]+distances[2]),2)*(full_length-(distances[0]+distances[1]+distances[2]))/(math.pow(full_length,2)))+00000+(my_load[3]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]))/(math.pow(full_length,2)))+00000+(my_load[4]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]))/(math.pow(full_length,2)))+00000+(my_load[5]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]))/(math.pow(full_length,2)))+00000+(my_load[6]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]))/(math.pow(full_length,2)))+00000+(my_load[7]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]))/(math.pow(full_length,2)))+00000+(my_load[8]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]))/(math.pow(full_length,2)))+00000+(my_load[9]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]))/(math.pow(full_length,2)))+00000+(my_load[10]*math.pow((distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]+distances[10]),2)*(full_length-(distances[0]+distances[1]+distances[2]+distances[3]+distances[4]+distances[5]+distances[6]+distances[7]+distances[8]+distances[9]+distances[10]))/(math.pow(full_length,2))))
                
                
                
                '''lengths_pointer=open("span_lengths.txt","a")    
                for i in all_length:
                    lengths_pointer.write(str(i))
                    lengths_pointer.close()'''
                    
                
                    
                    
                print("length = ",full_length)
                print("omega = ",omega)
                
                fem_pointer=open("fixed_end_moments.txt","a")
                fem_pointer.write(str(indice)+"\n\t"+str(fem1)+"\n\t"+str(fem2)+"\n")
                fem_pointer.close()
                
                    
                fem_sheet.cell(column=1,row=2*indice+1,value=fem1)
                fem_sheet.cell(column=1,row=2*indice+2,value=fem2)
                
                
                wb.save('all_datos.xlsx')
                
                    
                
                
                
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                    
                print(fem1)
                print(fem2)
                del distances
                
                
 ############fn6 ends 
                
            global indexer
            indexer=0
            for i in range(0,int(span)):
                incc=i
                bttn=Button(second_frame,text="load",width=5,command=fn6(indexer,indexer+int(cum_load_dist[i].get()),all_length[i],float(distributed_loads[i].get()),incc))
                bttn.place(x=257.5+515*i,y=130)
                indexer=indexer+int(cum_load_dist[i].get())
            
                        
#
    
    
#
        
                       
               
###########fn 2 =1 tabb    fn4=2 tabb   fn 5=3tabb    fn6=4tabb                
        def this_is_calculation(decesion1,decesion2):
            
            
                                    
            
            def left_fixed_right_hinge():
                wb=load_workbook('all_datos.xlsx')
                inertia_sheet=wb['inertias']
                span_sheet=wb['span_lengths']
                fem_sheet=wb['fixed_end_moments']
    
                L=[]
                I=[]
                col_inertia=inertia_sheet['a']
                span=len(col_inertia)
                col_len=span_sheet['a']
                
                for i in range(0,len(inertia_sheet['a'])):
                    I.append(col_inertia[i].value)
                    
                
                for i in range(0,len(inertia_sheet['a'])):
                    
                    L.append(col_len[i].value)
                    
                for i in range(0,100):
                    L.append(99)
                    I.append(99)
                
                
                
                
                
                
                mother=sy.Matrix([
                                [4*I[0]/L[0]+4*I[1]/L[1],2*I[1]/L[1],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [2*I[1]/L[1],4*I[1]/L[1]+4*I[2]/L[2],2*I[2]/L[2],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,2*I[2]/L[2],4*I[2]/L[2]+4*I[3]/L[3],2*I[3]/L[3],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,2*I[3]/L[3],4*I[3]/L[3]+4*I[4]/L[4],2*I[4]/L[4],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,2*I[4]/L[4],4*I[4]/L[4]+4*I[5]/L[5],2*I[5]/L[5],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,2*I[5]/L[5],4*I[5]/L[5]+4*I[6]/L[6],2*I[6]/L[6],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,2*I[6]/L[6],4*I[6]/L[6]+4*I[7]/L[7],2*I[7]/L[7],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,2*I[7]/L[7],4*I[7]/L[7]+4*I[8]/L[8],2*I[8]/L[8],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,2*I[8]/L[8],4*I[8]/L[8]+4*I[9]/L[9],2*I[9]/L[9],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,2*I[9]/L[9],4*I[9]/L[9]+4*I[10]/L[10],2*I[10]/L[10],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,2*I[10]/L[10],4*I[10]/L[10]+4*I[11]/L[11],2*I[11]/L[11],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,2*I[11]/L[11],4*I[11]/L[11]+4*I[12]/L[12],2*I[12]/L[12],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,2*I[12]/L[12],4*I[12]/L[12]+4*I[13]/L[13],2*I[13]/L[13],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,2*I[13]/L[13],4*I[13]/L[13]+4*I[14]/L[14],2*I[14]/L[14],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[14]/L[14],4*I[14]/L[14]+4*I[15]/L[15],2*I[15]/L[15],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[15]/L[15],4*I[15]/L[15]+4*I[16]/L[16],2*I[16]/L[16],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[16]/L[16],4*I[16]/L[16]+4*I[17]/L[17],2*I[17]/L[17],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[17]/L[17],4*I[17]/L[17]+4*I[18]/L[18],2*I[18]/L[18],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[18]/L[18],4*I[18]/L[18]+4*I[19]/L[19],2*I[19]/L[19],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[19]/L[19],4*I[19]/L[19]+4*I[20]/L[20],2*I[20]/L[20],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[20]/L[20],4*I[20]/L[20]+4*I[21]/L[21],2*I[21]/L[21],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[21]/L[21],4*I[21]/L[21]+4*I[22]/L[22],2*I[22]/L[22],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[22]/L[22],4*I[22]/L[22]+4*I[23]/L[23],2*I[23]/L[23],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[23]/L[23],4*I[23]/L[23]+4*I[24]/L[24],2*I[24]/L[24],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[24]/L[24],4*I[24]/L[24]+4*I[25]/L[25],2*I[25]/L[25],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[25]/L[25],4*I[25]/L[25]+4*I[26]/L[26],2*I[26]/L[26],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[26]/L[26],4*I[26]/L[26]+4*I[27]/L[27],2*I[27]/L[27],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[27]/L[27],4*I[27]/L[27]+4*I[28]/L[28],2*I[28]/L[28],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[28]/L[28],4*I[28]/L[28]+4*I[29]/L[29],2*I[29]/L[29],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[29]/L[29],4*I[29]/L[29]+4*I[30]/L[30],2*I[30]/L[30],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[30]/L[30],4*I[30]/L[30]+4*I[31]/L[31],2*I[31]/L[31],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[31]/L[31],4*I[31]/L[31]+4*I[32]/L[32],2*I[32]/L[32],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[32]/L[32],4*I[32]/L[32]+4*I[33]/L[33],2*I[33]/L[33],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[33]/L[33],4*I[33]/L[33]+4*I[34]/L[34],2*I[34]/L[34],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[34]/L[34],4*I[34]/L[34]+4*I[35]/L[35],2*I[35]/L[35],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[35]/L[35],4*I[35]/L[35]+4*I[36]/L[36],2*I[36]/L[36],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[36]/L[36],4*I[36]/L[36]+4*I[37]/L[37],2*I[37]/L[37],0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[37]/L[37],4*I[37]/L[37]+4*I[38]/L[38],2*I[38]/L[38],0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[38]/L[38],4*I[38]/L[38]+4*I[39]/L[39],2*I[39]/L[39],0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[39]/L[39],4*I[39]/L[39]+4*I[40]/L[40],2*I[40]/L[40],0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[40]/L[40],4*I[40]/L[40]+4*I[41]/L[41],2*I[41]/L[41],0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[41]/L[41],4*I[41]/L[41]+4*I[42]/L[42],2*I[42]/L[42],0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[42]/L[42],4*I[42]/L[42]+4*I[43]/L[43],2*I[43]/L[43],0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[43]/L[43],4*I[43]/L[43]+4*I[44]/L[44],2*I[44]/L[44],0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[44]/L[44],4*I[44]/L[44]+4*I[45]/L[45],2*I[45]/L[45],0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[45]/L[45],4*I[45]/L[45]+4*I[46]/L[46],2*I[46]/L[46],0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[46]/L[46],4*I[46]/L[46]+4*I[47]/L[47],2*I[47]/L[47],0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[47]/L[47],4*I[47]/L[47]+4*I[48]/L[48],2*I[48]/L[48],0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[48]/L[48],4*I[48]/L[48]+4*I[49]/L[49],2*I[49]/L[49],0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[49]/L[49],4*I[49]/L[49]+4*I[50]/L[50],2*I[50]/L[50],0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[50]/L[50],4*I[50]/L[50]+4*I[51]/L[51],2*I[51]/L[51]]
                    
                    
                    
                    
                    
                                ])
                
                
                
                                                                                                                                                                                                                                                                                            ####
                
                  
                
                
                order1=span-1
                for i in range (order1,51):
                    mother.row_del(order1)
                    mother.col_del(order1+1)
                    
                    
                '''
                no_of_zeros=mother.shape[1]-2
                first_eqn=zeros(1, mother.shape[1])
                first_eqn[0]=4*I[0]/L[0]
                first_eqn[1]=2*I[0]/L[0]
                
                '''
                
                
                
                
                
                
                last_eqn=zeros(1, mother.shape[1])
                last_eqn[mother.shape[1]-1]=4*I[span-1]/L[span-1]
                last_eqn[mother.shape[1]-2]=2*I[span-1]/L[span-1]
                   
                '''x=np.vstack([first_eqn,mother])'''
                final_mat=np.vstack([mother,last_eqn])
                   
                
                    
                
                
                last_fem_serial=len(fem_sheet['a'])    
                
                
                
                
                sec_lim=(last_fem_serial-2)/2
                fem_for_mid_eqn=[]
                
                a=2
                for i in range(0,int(sec_lim)):
                    gg=-1*(fem_sheet.cell(column=1,row=a).value+fem_sheet.cell(column=1,row=a+1).value)
                    fem_for_mid_eqn.append(gg)
                    a=a+2
                fem_for_mid_eqn.append(-1*fem_sheet.cell(column=1,row=last_fem_serial).value)
                constant2_matrix=sy.Matrix([fem_for_mid_eqn])
                constant_matrix=constant2_matrix.transpose()
                
                print("\n\n")
                
                
                ################
                
                
                final_mat_under_sympy=sy.Matrix(final_mat)
                
                
                
                solution=final_mat_under_sympy.inv()*constant_matrix
                
                
                for i in solution:
                    print(i)
                
                soln=[]
                soln.append(0)
                for i in solution:
                    soln.append(i)
                print("\n\n")
                for i in soln:
                    print(i)
                    
                    
                fem=[]
                for i in range(1,last_fem_serial+1):
                    fem.append(fem_sheet.cell(column=1,row=i).value)
                
                
                
                thetas=len(soln)-1
                
                z=0
                moments_final=[]
                
                
                print("\n\n")
                
                
                
                
                for i in range(0,thetas):
                    qq=fem[z]+4*I[i]*soln[i]/L[i]+2*I[i]*soln[i+1]/L[i]
                    moments_final.append(qq)
                    rr=fem[z+1]+2*I[i]*soln[i]/L[i]+4*I[i]*soln[i+1]/L[i]
                    moments_final.append(rr)
                    z=z+2
                
                for i in moments_final:
                    print(i)
                
                
                
                moments_sheet=wb.create_sheet('moments_after_calculation')
                
                for i in range(1,len(moments_final)+1):
                    moments_sheet.cell(column=1,row=i,value=str(moments_final[i-1]))
                 
                wb.save('all_datos.xlsx')
                
                            
#
            def both_end_fixed():
                wb=load_workbook('all_datos.xlsx')
                inertia_sheet=wb['inertias']
                span_sheet=wb['span_lengths']
                fem_sheet=wb['fixed_end_moments']
    
                L=[]
                I=[]
                col_inertia=inertia_sheet['a']
                span=len(col_inertia)
                col_len=span_sheet['a']
                
                for i in range(0,len(inertia_sheet['a'])):
                    I.append(col_inertia[i].value)
                    
                
                for i in range(0,len(inertia_sheet['a'])):
                    
                    L.append(col_len[i].value)
                    
                for i in range(0,100):
                    L.append(99)
                    I.append(99)
                
                
                
                
                
                
                mother=sy.Matrix([
                                [4*I[0]/L[0]+4*I[1]/L[1],2*I[1]/L[1],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [2*I[1]/L[1],4*I[1]/L[1]+4*I[2]/L[2],2*I[2]/L[2],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,2*I[2]/L[2],4*I[2]/L[2]+4*I[3]/L[3],2*I[3]/L[3],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,2*I[3]/L[3],4*I[3]/L[3]+4*I[4]/L[4],2*I[4]/L[4],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,2*I[4]/L[4],4*I[4]/L[4]+4*I[5]/L[5],2*I[5]/L[5],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,2*I[5]/L[5],4*I[5]/L[5]+4*I[6]/L[6],2*I[6]/L[6],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,2*I[6]/L[6],4*I[6]/L[6]+4*I[7]/L[7],2*I[7]/L[7],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,2*I[7]/L[7],4*I[7]/L[7]+4*I[8]/L[8],2*I[8]/L[8],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,2*I[8]/L[8],4*I[8]/L[8]+4*I[9]/L[9],2*I[9]/L[9],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,2*I[9]/L[9],4*I[9]/L[9]+4*I[10]/L[10],2*I[10]/L[10],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,2*I[10]/L[10],4*I[10]/L[10]+4*I[11]/L[11],2*I[11]/L[11],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,2*I[11]/L[11],4*I[11]/L[11]+4*I[12]/L[12],2*I[12]/L[12],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,2*I[12]/L[12],4*I[12]/L[12]+4*I[13]/L[13],2*I[13]/L[13],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,2*I[13]/L[13],4*I[13]/L[13]+4*I[14]/L[14],2*I[14]/L[14],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[14]/L[14],4*I[14]/L[14]+4*I[15]/L[15],2*I[15]/L[15],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[15]/L[15],4*I[15]/L[15]+4*I[16]/L[16],2*I[16]/L[16],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[16]/L[16],4*I[16]/L[16]+4*I[17]/L[17],2*I[17]/L[17],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[17]/L[17],4*I[17]/L[17]+4*I[18]/L[18],2*I[18]/L[18],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[18]/L[18],4*I[18]/L[18]+4*I[19]/L[19],2*I[19]/L[19],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[19]/L[19],4*I[19]/L[19]+4*I[20]/L[20],2*I[20]/L[20],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[20]/L[20],4*I[20]/L[20]+4*I[21]/L[21],2*I[21]/L[21],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[21]/L[21],4*I[21]/L[21]+4*I[22]/L[22],2*I[22]/L[22],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[22]/L[22],4*I[22]/L[22]+4*I[23]/L[23],2*I[23]/L[23],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[23]/L[23],4*I[23]/L[23]+4*I[24]/L[24],2*I[24]/L[24],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[24]/L[24],4*I[24]/L[24]+4*I[25]/L[25],2*I[25]/L[25],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[25]/L[25],4*I[25]/L[25]+4*I[26]/L[26],2*I[26]/L[26],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[26]/L[26],4*I[26]/L[26]+4*I[27]/L[27],2*I[27]/L[27],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[27]/L[27],4*I[27]/L[27]+4*I[28]/L[28],2*I[28]/L[28],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[28]/L[28],4*I[28]/L[28]+4*I[29]/L[29],2*I[29]/L[29],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[29]/L[29],4*I[29]/L[29]+4*I[30]/L[30],2*I[30]/L[30],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[30]/L[30],4*I[30]/L[30]+4*I[31]/L[31],2*I[31]/L[31],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[31]/L[31],4*I[31]/L[31]+4*I[32]/L[32],2*I[32]/L[32],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[32]/L[32],4*I[32]/L[32]+4*I[33]/L[33],2*I[33]/L[33],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[33]/L[33],4*I[33]/L[33]+4*I[34]/L[34],2*I[34]/L[34],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[34]/L[34],4*I[34]/L[34]+4*I[35]/L[35],2*I[35]/L[35],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[35]/L[35],4*I[35]/L[35]+4*I[36]/L[36],2*I[36]/L[36],0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[36]/L[36],4*I[36]/L[36]+4*I[37]/L[37],2*I[37]/L[37],0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[37]/L[37],4*I[37]/L[37]+4*I[38]/L[38],2*I[38]/L[38],0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[38]/L[38],4*I[38]/L[38]+4*I[39]/L[39],2*I[39]/L[39],0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[39]/L[39],4*I[39]/L[39]+4*I[40]/L[40],2*I[40]/L[40],0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[40]/L[40],4*I[40]/L[40]+4*I[41]/L[41],2*I[41]/L[41],0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[41]/L[41],4*I[41]/L[41]+4*I[42]/L[42],2*I[42]/L[42],0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[42]/L[42],4*I[42]/L[42]+4*I[43]/L[43],2*I[43]/L[43],0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[43]/L[43],4*I[43]/L[43]+4*I[44]/L[44],2*I[44]/L[44],0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[44]/L[44],4*I[44]/L[44]+4*I[45]/L[45],2*I[45]/L[45],0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[45]/L[45],4*I[45]/L[45]+4*I[46]/L[46],2*I[46]/L[46],0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[46]/L[46],4*I[46]/L[46]+4*I[47]/L[47],2*I[47]/L[47],0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[47]/L[47],4*I[47]/L[47]+4*I[48]/L[48],2*I[48]/L[48],0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[48]/L[48],4*I[48]/L[48]+4*I[49]/L[49],2*I[49]/L[49],0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[49]/L[49],4*I[49]/L[49]+4*I[50]/L[50],2*I[50]/L[50]],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[50]/L[50],4*I[50]/L[50]+4*I[51]/L[51]]
                    
                    
                    
                    
                    
                                ])
                
                
                
                                                                                                                                                                                                                                                                                            ####
                
                  
                
                
                order1=span-1
                for i in range (order1,51):
                    mother.row_del(order1)
                    mother.col_del(order1)
                
                
                
                last_fem_serial=len(fem_sheet['a'])    
                
                
                #first_fem=-1*fem_sheet['a1'].value
                
                sec_lim=(last_fem_serial-2)/2
                fem_for_mid_eqn=[]
                #fem_for_mid_eqn.append(first_fem)
                a=2
                for i in range(0,int(sec_lim)):
                    gg=-1*(fem_sheet.cell(column=1,row=a).value+fem_sheet.cell(column=1,row=a+1).value)
                    fem_for_mid_eqn.append(gg)
                    a=a+2
                #fem_for_mid_eqn.append(-1*fem_sheet.cell(column=1,row=last_fem_serial).value)
                constant2_matrix=sy.Matrix([fem_for_mid_eqn])
                constant_matrix=constant2_matrix.transpose()
                
                ################
                
                final_mat_under_sympy=sy.Matrix(mother)
                
                
                
                solution=final_mat_under_sympy.inv()*constant_matrix
                
                soln=[]
                soln.append(0)
                for i in solution:
                    soln.append(i)
                soln.append(0)
                
                print("\n\n")
                
                #print(soln)
                
                
                fem=[]
                for i in range(1,last_fem_serial+1):
                    fem.append(fem_sheet.cell(column=1,row=i).value)
                
                
                thetas=len(soln)-1
                
                
                z=0
                moments_final=[]
                
                
                
                
                for i in range(0,thetas):
                    qq=fem[z]+4*I[i]*soln[i]/L[i]+2*I[i]*soln[i+1]/L[i]
                    moments_final.append(qq)
                    rr=fem[z+1]+2*I[i]*soln[i]/L[i]+4*I[i]*soln[i+1]/L[i]
                    moments_final.append(rr)
                    z=z+2
                
                moments_sheet=wb.create_sheet('moments_after_calculation')
                
                for i in range(1,len(moments_final)+1):
                    moments_sheet.cell(column=1,row=i,value=str(moments_final[i-1]))
                 
                wb.save('all_datos.xlsx')
                
                
                
                
                
            




#

            def both_end_hinge_caln():
                wb=load_workbook('all_datos.xlsx')
                inertia_sheet=wb['inertias']
                span_sheet=wb['span_lengths']
                fem_sheet=wb['fixed_end_moments']
                
                L=[]
                I=[]
                col_inertia=inertia_sheet['a']
                span=len(col_inertia)
                col_len=span_sheet['a']
                
                for i in range(0,len(inertia_sheet['a'])):
                    I.append(col_inertia[i].value)
                    
                
                for i in range(0,len(inertia_sheet['a'])):
                    
                    L.append(col_len[i].value)
                    
                for i in range(0,100):
                    L.append(99)
                    I.append(99)
                
                
                
                
                
                
                mother=sy.Matrix([
                                [2*I[0]/L[0],4*I[0]/L[0]+4*I[1]/L[1],2*I[1]/L[1],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,2*I[1]/L[1],4*I[1]/L[1]+4*I[2]/L[2],2*I[2]/L[2],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,2*I[2]/L[2],4*I[2]/L[2]+4*I[3]/L[3],2*I[3]/L[3],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,2*I[3]/L[3],4*I[3]/L[3]+4*I[4]/L[4],2*I[4]/L[4],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,2*I[4]/L[4],4*I[4]/L[4]+4*I[5]/L[5],2*I[5]/L[5],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,2*I[5]/L[5],4*I[5]/L[5]+4*I[6]/L[6],2*I[6]/L[6],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,2*I[6]/L[6],4*I[6]/L[6]+4*I[7]/L[7],2*I[7]/L[7],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,2*I[7]/L[7],4*I[7]/L[7]+4*I[8]/L[8],2*I[8]/L[8],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,2*I[8]/L[8],4*I[8]/L[8]+4*I[9]/L[9],2*I[9]/L[9],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,2*I[9]/L[9],4*I[9]/L[9]+4*I[10]/L[10],2*I[10]/L[10],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,2*I[10]/L[10],4*I[10]/L[10]+4*I[11]/L[11],2*I[11]/L[11],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,2*I[11]/L[11],4*I[11]/L[11]+4*I[12]/L[12],2*I[12]/L[12],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,2*I[12]/L[12],4*I[12]/L[12]+4*I[13]/L[13],2*I[13]/L[13],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[13]/L[13],4*I[13]/L[13]+4*I[14]/L[14],2*I[14]/L[14],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[14]/L[14],4*I[14]/L[14]+4*I[15]/L[15],2*I[15]/L[15],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[15]/L[15],4*I[15]/L[15]+4*I[16]/L[16],2*I[16]/L[16],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[16]/L[16],4*I[16]/L[16]+4*I[17]/L[17],2*I[17]/L[17],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[17]/L[17],4*I[17]/L[17]+4*I[18]/L[18],2*I[18]/L[18],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[18]/L[18],4*I[18]/L[18]+4*I[19]/L[19],2*I[19]/L[19],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[19]/L[19],4*I[19]/L[19]+4*I[20]/L[20],2*I[20]/L[20],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[20]/L[20],4*I[20]/L[20]+4*I[21]/L[21],2*I[21]/L[21],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[21]/L[21],4*I[21]/L[21]+4*I[22]/L[22],2*I[22]/L[22],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[22]/L[22],4*I[22]/L[22]+4*I[23]/L[23],2*I[23]/L[23],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[23]/L[23],4*I[23]/L[23]+4*I[24]/L[24],2*I[24]/L[24],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[24]/L[24],4*I[24]/L[24]+4*I[25]/L[25],2*I[25]/L[25],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[25]/L[25],4*I[25]/L[25]+4*I[26]/L[26],2*I[26]/L[26],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[26]/L[26],4*I[26]/L[26]+4*I[27]/L[27],2*I[27]/L[27],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[27]/L[27],4*I[27]/L[27]+4*I[28]/L[28],2*I[28]/L[28],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[28]/L[28],4*I[28]/L[28]+4*I[29]/L[29],2*I[29]/L[29],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[29]/L[29],4*I[29]/L[29]+4*I[30]/L[30],2*I[30]/L[30],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[30]/L[30],4*I[30]/L[30]+4*I[31]/L[31],2*I[31]/L[31],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[31]/L[31],4*I[31]/L[31]+4*I[32]/L[32],2*I[32]/L[32],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[32]/L[32],4*I[32]/L[32]+4*I[33]/L[33],2*I[33]/L[33],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[33]/L[33],4*I[33]/L[33]+4*I[34]/L[34],2*I[34]/L[34],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[34]/L[34],4*I[34]/L[34]+4*I[35]/L[35],2*I[35]/L[35],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[35]/L[35],4*I[35]/L[35]+4*I[36]/L[36],2*I[36]/L[36],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[36]/L[36],4*I[36]/L[36]+4*I[37]/L[37],2*I[37]/L[37],0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[37]/L[37],4*I[37]/L[37]+4*I[38]/L[38],2*I[38]/L[38],0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[38]/L[38],4*I[38]/L[38]+4*I[39]/L[39],2*I[39]/L[39],0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[39]/L[39],4*I[39]/L[39]+4*I[40]/L[40],2*I[40]/L[40],0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[40]/L[40],4*I[40]/L[40]+4*I[41]/L[41],2*I[41]/L[41],0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[41]/L[41],4*I[41]/L[41]+4*I[42]/L[42],2*I[42]/L[42],0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[42]/L[42],4*I[42]/L[42]+4*I[43]/L[43],2*I[43]/L[43],0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[43]/L[43],4*I[43]/L[43]+4*I[44]/L[44],2*I[44]/L[44],0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[44]/L[44],4*I[44]/L[44]+4*I[45]/L[45],2*I[45]/L[45],0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[45]/L[45],4*I[45]/L[45]+4*I[46]/L[46],2*I[46]/L[46],0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[46]/L[46],4*I[46]/L[46]+4*I[47]/L[47],2*I[47]/L[47],0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[47]/L[47],4*I[47]/L[47]+4*I[48]/L[48],2*I[48]/L[48],0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[48]/L[48],4*I[48]/L[48]+4*I[49]/L[49],2*I[49]/L[49],0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[49]/L[49],4*I[49]/L[49]+4*I[50]/L[50],2*I[50]/L[50],0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[50]/L[50],4*I[50]/L[50]+4*I[51]/L[51],2*I[51]/L[51]]
                    
                    
                    
                    
                    
                                ])
                
                
                
                                                                                                                                                                                                                                                                                            ####
                
                  
                
                
                order1=span-1
                for i in range (order1,51):
                    mother.row_del(order1)
                    mother.col_del(order1+2)
                
                no_of_zeros=mother.shape[1]-2
                first_eqn=zeros(1, mother.shape[1])
                first_eqn[0]=4*I[0]/L[0]
                first_eqn[1]=2*I[0]/L[0]
                
                last_eqn=zeros(1, mother.shape[1])
                last_eqn[mother.shape[1]-1]=4*I[span-1]/L[span-1]
                last_eqn[mother.shape[1]-2]=2*I[span-1]/L[span-1]
                   
                x=np.vstack([first_eqn,mother])
                final_mat=np.vstack([x,last_eqn])
                   
                
                    
                '''print(final_mat)'''
                
                last_fem_serial=len(fem_sheet['a'])    
                '''print(last_fem)'''
                
                first_fem=-1*fem_sheet['a1'].value
                '''print(first_fem)'''
                sec_lim=(last_fem_serial-2)/2
                fem_for_mid_eqn=[]
                fem_for_mid_eqn.append(first_fem)
                a=2
                for i in range(0,int(sec_lim)):
                    gg=-1*(fem_sheet.cell(column=1,row=a).value+fem_sheet.cell(column=1,row=a+1).value)
                    fem_for_mid_eqn.append(gg)
                    a=a+2
                fem_for_mid_eqn.append(-1*fem_sheet.cell(column=1,row=last_fem_serial).value)
                constant2_matrix=sy.Matrix([fem_for_mid_eqn])
                constant_matrix=constant2_matrix.transpose()
                '''print(constant_matrix)'''
                ################
                
                final_mat_under_sympy=sy.Matrix(final_mat)
                
                
                
                solution=final_mat_under_sympy.inv()*constant_matrix
                
                ''''ccc=bbb.slove(aaa)'''
                for i in solution:
                    print(i)
                
                
                print("\n\n")
                
                fem=[]
                for i in range(1,last_fem_serial+1):
                    fem.append(fem_sheet.cell(column=1,row=i).value)
                
                
                thetas=len(solution)-1
                
                
                z=0
                moments_final=[]
                
                
                
                
                for i in range(0,thetas):
                    qq=fem[z]+4*I[i]*solution[i]/L[i]+2*I[i]*solution[i+1]/L[i]
                    moments_final.append(qq)
                    rr=fem[z+1]+2*I[i]*solution[i]/L[i]+4*I[i]*solution[i+1]/L[i]
                    moments_final.append(rr)
                    z=z+2
                
                
                
                
                moments_sheet=wb.create_sheet('moments_after_calculation')
                
                for i in range(1,len(moments_final)+1):
                    moments_sheet.cell(column=1,row=i,value=str(moments_final[i-1]))
                 
                wb.save('all_datos.xlsx')
                



#

            def left_hinged_right_fixed():
                wb=load_workbook('all_datos.xlsx')
                inertia_sheet=wb['inertias']
                span_sheet=wb['span_lengths']
                fem_sheet=wb['fixed_end_moments']
                
                L=[]
                I=[]
                col_inertia=inertia_sheet['a']
                span=len(col_inertia)
                col_len=span_sheet['a']
                
                for i in range(0,len(inertia_sheet['a'])):
                    I.append(col_inertia[i].value)
                    
                
                for i in range(0,len(inertia_sheet['a'])):
                    
                    L.append(col_len[i].value)
                    
                for i in range(0,100):
                    L.append(99)
                    I.append(99)
                
                
                
                
                
                
                mother=sy.Matrix([
                                [2*I[0]/L[0],4*I[0]/L[0]+4*I[1]/L[1],2*I[1]/L[1],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,2*I[1]/L[1],4*I[1]/L[1]+4*I[2]/L[2],2*I[2]/L[2],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,2*I[2]/L[2],4*I[2]/L[2]+4*I[3]/L[3],2*I[3]/L[3],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,2*I[3]/L[3],4*I[3]/L[3]+4*I[4]/L[4],2*I[4]/L[4],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,2*I[4]/L[4],4*I[4]/L[4]+4*I[5]/L[5],2*I[5]/L[5],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,2*I[5]/L[5],4*I[5]/L[5]+4*I[6]/L[6],2*I[6]/L[6],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,2*I[6]/L[6],4*I[6]/L[6]+4*I[7]/L[7],2*I[7]/L[7],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,2*I[7]/L[7],4*I[7]/L[7]+4*I[8]/L[8],2*I[8]/L[8],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,2*I[8]/L[8],4*I[8]/L[8]+4*I[9]/L[9],2*I[9]/L[9],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,2*I[9]/L[9],4*I[9]/L[9]+4*I[10]/L[10],2*I[10]/L[10],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,2*I[10]/L[10],4*I[10]/L[10]+4*I[11]/L[11],2*I[11]/L[11],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,2*I[11]/L[11],4*I[11]/L[11]+4*I[12]/L[12],2*I[12]/L[12],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,2*I[12]/L[12],4*I[12]/L[12]+4*I[13]/L[13],2*I[13]/L[13],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[13]/L[13],4*I[13]/L[13]+4*I[14]/L[14],2*I[14]/L[14],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[14]/L[14],4*I[14]/L[14]+4*I[15]/L[15],2*I[15]/L[15],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[15]/L[15],4*I[15]/L[15]+4*I[16]/L[16],2*I[16]/L[16],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[16]/L[16],4*I[16]/L[16]+4*I[17]/L[17],2*I[17]/L[17],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[17]/L[17],4*I[17]/L[17]+4*I[18]/L[18],2*I[18]/L[18],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[18]/L[18],4*I[18]/L[18]+4*I[19]/L[19],2*I[19]/L[19],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[19]/L[19],4*I[19]/L[19]+4*I[20]/L[20],2*I[20]/L[20],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[20]/L[20],4*I[20]/L[20]+4*I[21]/L[21],2*I[21]/L[21],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[21]/L[21],4*I[21]/L[21]+4*I[22]/L[22],2*I[22]/L[22],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[22]/L[22],4*I[22]/L[22]+4*I[23]/L[23],2*I[23]/L[23],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[23]/L[23],4*I[23]/L[23]+4*I[24]/L[24],2*I[24]/L[24],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[24]/L[24],4*I[24]/L[24]+4*I[25]/L[25],2*I[25]/L[25],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[25]/L[25],4*I[25]/L[25]+4*I[26]/L[26],2*I[26]/L[26],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[26]/L[26],4*I[26]/L[26]+4*I[27]/L[27],2*I[27]/L[27],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[27]/L[27],4*I[27]/L[27]+4*I[28]/L[28],2*I[28]/L[28],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[28]/L[28],4*I[28]/L[28]+4*I[29]/L[29],2*I[29]/L[29],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[29]/L[29],4*I[29]/L[29]+4*I[30]/L[30],2*I[30]/L[30],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[30]/L[30],4*I[30]/L[30]+4*I[31]/L[31],2*I[31]/L[31],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[31]/L[31],4*I[31]/L[31]+4*I[32]/L[32],2*I[32]/L[32],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[32]/L[32],4*I[32]/L[32]+4*I[33]/L[33],2*I[33]/L[33],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[33]/L[33],4*I[33]/L[33]+4*I[34]/L[34],2*I[34]/L[34],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[34]/L[34],4*I[34]/L[34]+4*I[35]/L[35],2*I[35]/L[35],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[35]/L[35],4*I[35]/L[35]+4*I[36]/L[36],2*I[36]/L[36],0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[36]/L[36],4*I[36]/L[36]+4*I[37]/L[37],2*I[37]/L[37],0,0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[37]/L[37],4*I[37]/L[37]+4*I[38]/L[38],2*I[38]/L[38],0,0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[38]/L[38],4*I[38]/L[38]+4*I[39]/L[39],2*I[39]/L[39],0,0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[39]/L[39],4*I[39]/L[39]+4*I[40]/L[40],2*I[40]/L[40],0,0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[40]/L[40],4*I[40]/L[40]+4*I[41]/L[41],2*I[41]/L[41],0,0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[41]/L[41],4*I[41]/L[41]+4*I[42]/L[42],2*I[42]/L[42],0,0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[42]/L[42],4*I[42]/L[42]+4*I[43]/L[43],2*I[43]/L[43],0,0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[43]/L[43],4*I[43]/L[43]+4*I[44]/L[44],2*I[44]/L[44],0,0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[44]/L[44],4*I[44]/L[44]+4*I[45]/L[45],2*I[45]/L[45],0,0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[45]/L[45],4*I[45]/L[45]+4*I[46]/L[46],2*I[46]/L[46],0,0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[46]/L[46],4*I[46]/L[46]+4*I[47]/L[47],2*I[47]/L[47],0,0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[47]/L[47],4*I[47]/L[47]+4*I[48]/L[48],2*I[48]/L[48],0,0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[48]/L[48],4*I[48]/L[48]+4*I[49]/L[49],2*I[49]/L[49],0],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[49]/L[49],4*I[49]/L[49]+4*I[50]/L[50],2*I[50]/L[50]],
                [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,2*I[50]/L[50],4*I[50]/L[50]+4*I[51]/L[51]]
                    
                    
                    
                    
                    
                                ])
                
                
                
                                                                                                                                                                                                                                                                                            ####
                
                  
                
                
                order1=span-1
                for i in range (order1,51):
                    mother.row_del(order1)
                    mother.col_del(order1+1)
                
                ##
                
                
                '''no_of_zeros=mother.shape[1]-2'''
                first_eqn=zeros(1, mother.shape[1])
                first_eqn[0]=4*I[0]/L[0]
                first_eqn[1]=2*I[0]/L[0]
                
                
                   
                final_mat=np.vstack([first_eqn,mother])
                
                
                
                #######lhs ok
                
                
                
                last_fem_serial=len(fem_sheet['a'])    
                
                
                first_fem=-1*fem_sheet['a1'].value
                
                sec_lim=(last_fem_serial-2)/2
                fem_for_mid_eqn=[]
                fem_for_mid_eqn.append(first_fem)
                a=2
                for i in range(0,int(sec_lim)):
                    gg=-1*(fem_sheet.cell(column=1,row=a).value+fem_sheet.cell(column=1,row=a+1).value)
                    fem_for_mid_eqn.append(gg)
                    a=a+2
                
                constant2_matrix=sy.Matrix([fem_for_mid_eqn])
                
                constant_matrix=constant2_matrix.transpose()
                
                print(constant_matrix)
                
                
                
                ################
                
                final_mat_under_sympy=sy.Matrix(final_mat)
                
                
                
                solution=final_mat_under_sympy.inv()*constant_matrix
                
                
                
                
                
                print("\n\n")
                
                fem=[]
                for i in range(1,last_fem_serial+1):
                    fem.append(fem_sheet.cell(column=1,row=i).value)
                
                
                soln=[]
                
                for i in solution:
                    soln.append(i)
                soln.append(0)
                
                
                for i in soln:
                    print(i)
                    
                print("\n\n")
                
                
                    
                fem=[]
                for i in range(1,last_fem_serial+1):
                    fem.append(fem_sheet.cell(column=1,row=i).value)
                
                
                
                thetas=len(soln)-1
                
                z=0
                moments_final=[]
                
                
                print("\n\n")
                
                
                
                
                for i in range(0,thetas):
                    qq=fem[z]+4*I[i]*soln[i]/L[i]+2*I[i]*soln[i+1]/L[i]
                    moments_final.append(qq)
                    rr=fem[z+1]+2*I[i]*soln[i]/L[i]+4*I[i]*soln[i+1]/L[i]
                    moments_final.append(rr)
                    z=z+2
                
                for i in moments_final:
                    print(i)
                
                moments_sheet=wb.create_sheet('moments_after_calculation')
                
                for i in range(1,len(moments_final)+1):
                    moments_sheet.cell(column=1,row=i,value=str(moments_final[i-1]))
                 
                wb.save('all_datos.xlsx')

#


            
            
                        
                        
            
            
            ####---------------
            if decesion1=="fixed" and (decesion2=="hinge" or decesion2=="roller"):
                left_fixed_right_hinge()
                
                
            if decesion2=="fixed" and (decesion1=="hinge" or decesion1=="roller"):
                left_hinged_right_fixed()
                
            if (decesion1=="hinge" or decesion1=="roller") and (decesion2=="hinge" or decesion2=="roller"):
                both_end_hinge_caln()
                
            if decesion1=="fixed" and decesion2=="fixed":
                
                both_end_fixed()
                
                
            
                
              
        def this_is_graph():
            def draw():
                wb=load_workbook('all_datos.xlsx')
                omega_sheet=wb['omegas']
                span_lengths_sheet=wb['span_lengths']
                moments_sheet=wb['moments_after_calculation']
                point_load_sheet=wb['poin_loads']
                distances_sheet=wb['distances_between_point_loads']
                
                
                total_spans=(len(span_lengths_sheet['a']))
                
                def fn1(span_number,b,cum_dist):
                    print("---------------------------------")
                    
                    p=[]
                    distances=[]
                    for i in range(1,100):
                        p.append(point_load_sheet.cell(row=i,column=span_number).value)
                        if point_load_sheet.cell(row=i+1,column=span_number).value==None:
                            break
                    for i in range(1,100):
                        distances.append(distances_sheet.cell(row=i,column=span_number).value)
                        if distances_sheet.cell(row=i+1,column=span_number).value==None:
                            break    
                    
                    print("\ndistances are")
                    for i in distances:
                        print(i)
                    distant=[]
                    xx=0
                    for i in range(0,len(p)):
                        xx=xx+distances[i]
                        distant.append(xx)
                        
                        
                        
                        
                        
                    print("\n point loads are")
                    for i in p:
                        print(i)
                    moment_left=-1*float((moments_sheet.cell(row=b,column=1).value))
                    moment_right=-1*float((moments_sheet.cell(row=b+1,column=1).value))
                    omega=omega_sheet.cell(row=span_number,column=1).value
                    L=span_lengths_sheet.cell(row=span_number,column=1).value
                    
                    print(f'moment left = {moment_left}  moment right = {moment_right}  omega= {omega}  span length= {L}')
                    
                     
                    print('\n')
                    arbit_sum=0
                    for i in range(0,len(p)):
                        arbit=p[i]*distant[i]
                        arbit_sum=arbit_sum+arbit
                    R_left=omega*L+sum(p)-(moment_left+moment_right+arbit_sum+omega*L**2/2)/L
                    print("R left is    = ",R_left)
                    R_right=(moment_left+moment_right+arbit_sum+omega*L**2/2)/L
                    ###############data points for sfd
                    
                    
                    
                    total_points=len(p)
                    print("total points  = ",total_points)
                    
                    point_y=[]
                    point_y.append(0)
                    point_y.append(R_left)
                    
                    
                    z=0
                    for i in range(0,total_points):
                        if z==0:
                            point_y.append(R_left-omega*distances[i])
                            point_y.append(point_y[2]-p[i])
                        else:
                            point_y.append(point_y[z+1]-omega*distances[i])
                            point_y.append(point_y[z+2]-p[i])
                        z+=2
                    point_y.append(-1*R_right)
                    point_y.append(0)
                    
                        
                    point_x=[]
                    point_x.append(0+cum_dist)
                    point_x.append(0+cum_dist)
                    for i in range(0,len(distant)):
                        point_x.append(cum_dist+distant[i])
                        point_x.append(cum_dist+distant[i])
                    point_x.append(L+cum_dist)
                    point_x.append(L+cum_dist)
                    
                    
                    for i in point_y:
                        print("y = ",i)
                        
                    for i in point_x:
                        print("x = ",i)
                    ##################
                    #moments points starts here
                    
                    
                    
                    
                    
                    print("\n\ndistant")
                    for i in distant:
                        print(i)
                        
                        
                    mnt_right=-1*moment_right
                    m_point_y=[]
                    
                    m_point_y.append(moment_left)
                    for i in range(1,L):
                        c=0
                        for j in distant:
                            if j<i:
                                c+=1
                        if c==0:
                            m_y=moment_left+R_left*i-(omega*i*i)/2
                            m_point_y.append(m_y)
                        else:
                            suml=0
                            for k in range(0,c):
                                suml=suml+(p[k])*(i-distant[k])
                            m_y=moment_left+R_left*i-(omega*i*i)/2-suml
                            m_point_y.append(m_y)
                    m_point_y.append(mnt_right)
                    
                    for i in m_point_y:
                        print(i)
                    
                    print("total then x es",len(m_point_y))
                    m_point_x=[]
                  
                    for i in range(0,len(m_point_y)):
                        m_point_x.append(i+cum_dist)
                    for i in m_point_x:
                        print(i)
                ####--------------------------     
                    for i in m_point_x:
                        moments_x_d.append(i)
                    
                    for i in m_point_y:
                        moments_y_d.append(i)
                    for i in point_x:
                        sfd_x_d.append(i)
                    for i in point_y:
                        sfd_y_d.append(i)
                   
                    
                    
                    
                    
                    
                    
                    
                    
                    #############
                    
                sfd_bmd_sheet=wb.create_sheet('drawings')   
                lengths=[]
                for i in range(1,len(span_lengths_sheet['a'])+1 ):
                    lengths.append(span_lengths_sheet.cell(row=i,column=1).value)
                    
                ###-----------
                moments_x_d=[]
                moments_y_d=[]
                sfd_x_d=[]
                sfd_y_d=[]
                
                ###-------------
                
                moments_x_d.append(0)
                moments_y_d.append(0)
                
                
                
                
                z=0
                bb=1
                su=0
                
                for i in range(1,total_spans+1):
                    fn1(i,bb,su)
                    su=su+lengths[i-1]
                    
                    
                    bb+=2
                
                
                sfd_x_d.append(0)
                sfd_y_d.append(0)
                
                
                sfd_bmd_sheet.cell(column=3,row=2+len(moments_x_d),value=moments_x_d[len(moments_x_d)-1])
                
                
                
                
                last_moment_x=max(moments_x_d)
                
                
                
                
                moments_x_d.append(last_moment_x)
                moments_y_d.append(0)
                
                moments_x_d.append(0)
                moments_y_d.append(0)
                
                
                
                
                
                
                
                
                ##########################
                
                sfd_bmd_sheet.cell(column=4,row=1,value=0)
                for i in range(0,len(moments_y_d)):
                    sfd_bmd_sheet.cell(column=4,row=i+2,value=moments_y_d[i])
                sfd_bmd_sheet.cell(column=4,row=2+len(moments_y_d),value=0)
                
                
                
                sfd_bmd_sheet.cell(column=3,row=1,value=0)
                for i in range(0,len(moments_x_d)):
                    sfd_bmd_sheet.cell(column=3,row=i+2,value=moments_x_d[i])
                sfd_bmd_sheet.cell(column=3,row=2+len(moments_x_d),value=moments_x_d[len(moments_x_d)-1])
                
                
                
                for i in range(0,len(sfd_x_d)):
                    sfd_bmd_sheet.cell(column=1,row=i+1,value=sfd_x_d[i])
                
                for i in range(0,len(sfd_y_d)):
                    sfd_bmd_sheet.cell(column=2,row=i+1,value=sfd_y_d[i])
                #############################
                
                
                fig,(sfd_plot,bmd_plot)=plt.subplots(2,1)
                
                
                sfd_plot.plot(sfd_x_d,sfd_y_d)
                bmd_plot.plot(moments_x_d,moments_y_d)
                
                sfd_plot.set_xlabel('SHEAR FORCE IN KN')
                bmd_plot.set_xlabel('BENDING MOMENT IN KN/unit')
                
                plt.show()
                
                
                
                
                
                
                
                
                
                
                    
                wb.save('all_datos_and_dwg.xlsx')
                
                
            
            draw()
        
        
        
        
        
        
        btncal=Button(second_frame,text="load data (first click this)",width=30,command=fn5)
        btncal.grid(row=8,column=0)
        
        btncal21=Button(second_frame,text="2nd calculate",width=30,command=lambda :this_is_calculation(left,right))
        btncal21.grid(row=8,column=1)
        
        btncal22=Button(second_frame,text="3rd show sfd bmd",width=30,command=this_is_graph)
        btncal22.grid(row=8,column=2)
        
        
        
        
###########fn 2 =1 tabb    fn4=2 tabb   fn 5=3tabb    fn6=4tabb      
        
        p_load=[]
        a=0
        n=0
        for i in range(0,len(cum_load_dist)):
           
            for j in range(0,int(cum_load_dist[i].get())):
                P_load=Entry(second_frame,width=5,bg="#99ccff")
                P_load.place(x=a+515/(int(cum_load_dist[i].get())+1),y=183)
                p_load.append(P_load)
                
                point_load_img=Label(second_frame,image=P)
                point_load_img.place(x=a+515/(int(cum_load_dist[i].get())+1),y=203)
                
                a=a+515/(int(cum_load_dist[i].get())+1)
                if j==int(cum_load_dist[i].get())-1:
                    n=n+1
                    a=515*n
                            
        only_dist=[]            
        b=0
        m=0            
        for i in range(0,len(cum_load_dist)):
            for j in range(0,int(cum_load_dist[i].get())+1):
                Only_dist=Entry(second_frame,borderwidth=4,width=3,bg="#99ccff")
                Only_dist.place(x=b+(515/(int(cum_load_dist[i].get())+1))/2*(2*j+1),y=315)
                only_dist.append(Only_dist)
                
                
                if j==int(cum_load_dist[i].get()):
                    m=m+1
                    b=515*m
                
            
            
###########fn 2 =1 tabb    fn4=2 tabb   fn 5=3tabb    fn6=4tabb        
        
        
        

    
    pl=Label(second_frame,height=10).grid(row=5,column=0,columnspan=int(span))
    
    
    cum_load_dist=[]
    list_in_mid_span=[]
    var_img=Variable
    in_spn=[]
    lbl=[]
    for i in range(0,int(span)):
        
        if i==0 and left=='fixed':
            
            lbl=Label(second_frame,image=img1,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)'''
            
            
        if i==0 and left=='hinge':
            lbl=Label(second_frame,image=img2,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)'''
        
        
        if i==0 and left=='roller':
            lbl=Label(second_frame,image=img3,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)'''
        
        
        if i==int(span)-1 and right=='fixed':
            lbl=Label(second_frame,image=img4,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)'''
            
            
        if i==int(span)-1 and right=='hinge':
            lbl=Label(second_frame,image=img5,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)''' 
            
            
        if i==int(span)-1 and right=='roller':
            lbl=Label(second_frame,image=img6,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)'''
        
        if i>0 and i!=int(span)-1:
            lbl=Label(second_frame,image=mid_img,borderwidth=0)
            lbl.grid(row=6,column=i)
            '''in_mid_span=Entry(second_frame,width=10)
            in_mid_span.grid(row=7,column=i)
            list_in_mid_span.append(in_mid_span)''' 
        Cum_load_dist=Entry(second_frame,width=40,bg="#99ccff")
        if i==0:
            Cum_load_dist.insert(0, 'number of point loads in this span(at least 1)')
            
        
        Cum_load_dist.grid(row=3,column=i)
        
        cum_load_dist.append(Cum_load_dist)
        if i==int(span)-1:
            pick_btn=Button(second_frame,text="ok",command=fn4)
            pick_btn.grid(row=4,column=i)
            
        
    distributed_loads=[]
    beam_width=[]
    beam_height=[]
        
    for i in range(0,int(span)):
        distributed_input=Entry(second_frame,width=5)
        distributed_input.place(x=257.5+515*i,y=275)
        distributed_loads.append(distributed_input)
        
        beam_w=Label(second_frame,text="B=",width=3)
        beam_w.place(x=257.5-100+515*i,y=340)
        
        
        b_width=Entry(second_frame,bg="#33FFA8",width=5)
        b_width.place(x=257.5-75+515*i,y=340)
        beam_width.append(b_width)
        
        beam_h=Label(second_frame,text="D=",width=3)
        beam_h.place(x=257.5+515*i,y=340)
        
        
        b_ht=Entry(second_frame,bg="#33FFA8",width=5)
        b_ht.place(x=257.5+25+515*i,y=340)
        beam_height.append(b_ht)
        
        
        
    
        
        
        
        
###########fn 2 =1 tabb    fn4=2 tabb   fn 5=3tabb    fn6=4tabb                    
        
       
    '''load_span_btn=Button(second_frame,text="input loads",width=20,command=fn4).grid(row=8,column=0)'''
    
    



def fn1():
    try:
        global span,left,right
        span=float(Ent1.get())
        
        left=(clicked.get())
        right=(clicked1.get())
        fn2()
        
    except ValueError:
        pass
    
    
            

label1=Label(second_frame,text="no. of span",width=10).grid(row=0,column=0)
Ent1=Entry(second_frame,width=10,borderwidth=5)

Ent1.grid(row=0,column=1)




clicked=StringVar()
clicked.set("select")

clicked1=StringVar()
clicked1.set("select")

label2=Label(second_frame,text="Left end",width=50).grid(row=1,column=0)

drop1=OptionMenu(second_frame,clicked,"fixed","hinge","roller").grid(row=1,column=1)

label3=Label(second_frame,text="Right end",width=50).grid(row=2,column=0)

drop2=OptionMenu(second_frame,clicked1,"fixed","hinge","roller").grid(row=2,column=1)

btn1=Button(second_frame,text="ok",width=30,command=fn1).grid(row=2,column=2)
'''btn2=Button(second_frame,text="ok",width=30,command=second_frame.close).grid(row=2,column=3)'''








    
    


















root.mainloop()
